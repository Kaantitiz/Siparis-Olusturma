import streamlit as st
import pandas as pd
import json
import io
import re
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

# Sayfa ayarları
st.set_page_config(
    page_title="BOSCH Sipariş İşlemleri",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Başlık
st.title("🏭 BOSCH Sipariş İşlemleri")
st.caption("3 Excel dosyasından son.json formatında çıktı oluşturma")

# Global değişkenler
if 'process_bosch' not in st.session_state:
    st.session_state.process_bosch = False
if 'final_df' not in st.session_state:
    st.session_state.final_df = None

def process_bosch_codes(bosch_ref):
    """Bosch ürün kodlarını işle - başına 3E- ekle ve boşlukları temizle"""
    if pd.isna(bosch_ref):
        return ''
    
    code_str = str(bosch_ref).strip()
    code_str = re.sub(r'\s+', '', code_str)  # Tüm boşlukları kaldır
    
    if not code_str.startswith('3E-'):
        code_str = '3E-' + code_str
    
    return code_str

def determine_depot_code(siparis_notu):
    """Sipariş Notu'ndan depo kodunu belirle - sadece belirli kodlar"""
    if pd.isna(siparis_notu) or siparis_notu == "":
        return ""
    
    siparis_str = str(siparis_notu).strip().upper()
    
    if len(siparis_str) >= 3:
        depo_kodu = siparis_str[:3]
        allowed_codes = ['AAS', 'DAS', 'MAS', 'BAS', 'EAS']
        return depo_kodu if depo_kodu in allowed_codes else ""
    
    return ""

def create_sutun1(siparis_notu, bosch_no, siparis_miktari=None, kalan_miktar=None):
    """Sütun1 oluştur - Sipariş Notu + Bosch No + (varsa miktar)"""
    siparis_str = str(siparis_notu) if pd.notna(siparis_notu) else ""
    bosch_str = str(bosch_no) if pd.notna(bosch_no) else ""
    
    siparis_clean = re.sub(r'\s+', '', siparis_str)
    bosch_clean = re.sub(r'\s+', '', bosch_str)
    
    base = f"{siparis_clean}{bosch_clean}"
    
    if siparis_miktari is not None and kalan_miktar is not None:
        try:
            if abs(float(siparis_miktari) - float(kalan_miktar)) < 0.001:
                return f"{base}_{int(kalan_miktar)}" if float(kalan_miktar).is_integer() else f"{base}_{float(kalan_miktar):.2f}"
        except (ValueError, TypeError):
            pass
    
    return base

def validate_dataframes(bakiye_df, inbound_df, siparis_df):
    """Yüklenen DataFrame'leri doğrula"""
    required_cols = {
        'bakiye': ['Sipariş Notu', 'Ürün Grubu', 'Bosch No', 'Fatura ve Sevk Edilmemiş Toplam'],
        'inbound': ['Cari', 'Sipariş No', 'Ürün Kodu', 'İrsaliye Miktarı'],
        'siparis': ['SIPARIS_NO', 'STOK_KODU', 'SIPARIS_MIKTARI', 'KALAN_MIKTAR']
    }
    
    errors = []
    
    for df_name, df, cols in zip(['bakiye', 'inbound', 'siparis'], 
                                [bakiye_df, inbound_df, siparis_df], 
                                required_cols.values()):
        missing = [col for col in cols if col not in df.columns]
        if missing:
            errors.append(f"{df_name} eksik kolonlar: {', '.join(missing)}")
    
    return errors if not errors else None

def process_bosch_three_excel(bakiye_raporu, inbound_excel, siparis_kalemleri):
    """BOSCH için 3-Excel işlemi - son.json formatında çıktı"""
    try:
        # 1. ADIM: Dosyaları yükle
        with st.spinner("📂 Dosyalar yükleniyor..."):
            bakiye_df = pd.read_excel(bakiye_raporu, engine='openpyxl')
            inbound_df = pd.read_excel(inbound_excel, engine='openpyxl')
            siparis_df = pd.read_excel(siparis_kalemleri, engine='openpyxl')
            
            # Veri doğrulama
            if errors := validate_dataframes(bakiye_df, inbound_df, siparis_df):
                st.error("⚠️ Veri doğrulama hataları:\n" + "\n".join(errors))
                return None

        # 2. ADIM: Bakiye verilerini işle
        with st.spinner("📊 Bakiye verileri işleniyor..."):
            bakiye_df['Bosch No'] = bakiye_df['Bosch No'].apply(process_bosch_codes)
            bakiye_df['Birleşik_Kod'] = (
                bakiye_df['Sipariş Notu'].astype(str).str.replace(' ', '') + 
                bakiye_df['Bosch No'].astype(str).str.replace(' ', '')
            )
            bakiye_df['Ürün Grubu'] = bakiye_df['Ürün Grubu'].replace({'TEDARİKÇİLER': 'TEDARİKÇİ'})

        # 3. ADIM: InBound verilerini ekle
        with st.spinner("📦 InBound verileri işleniyor..."):
            bosch_inbound = inbound_df[
                inbound_df['Cari'].astype(str).str.contains('BOSCH', case=False, na=False)
            ]
            
            if not bosch_inbound.empty:
                inbound_data = []
                for _, row in bosch_inbound.iterrows():
                    inbound_data.append({
                        'Sipariş Notu': row['Sipariş No'],
                        'Ürün Grubu': 'DEPO',
                        'Bosch No': process_bosch_codes(row['Ürün Kodu']),
                        'Fatura ve Sevk Edilmemiş Toplam': row['İrsaliye Miktarı'],
                        'Birleşik_Kod': str(row['Sipariş No']).replace(' ', '') + 
                                       process_bosch_codes(row['Ürün Kodu']).replace(' ', '')
                    })
                
                bakiye_df = pd.concat([bakiye_df, pd.DataFrame(inbound_data)], ignore_index=True)

        # 4. ADIM: Sipariş verilerini hazırla
        with st.spinner("📋 Sipariş verileri hazırlanıyor..."):
            siparis_df['Siparis_Birlesik'] = (
                siparis_df['SIPARIS_NO'].astype(str).str.replace(' ', '') + 
                siparis_df['STOK_KODU'].astype(str).str.replace(' ', '')
            )
            
            siparis_gruplu = siparis_df.groupby('Siparis_Birlesik').agg({
                'SIPARIS_MIKTARI': 'sum',
                'KALAN_MIKTAR': 'sum'
            }).reset_index()

        # 5. ADIM: Gelişmiş eşleştirme
        with st.spinner("🔍 Veriler eşleştiriliyor..."):
            processed_rows = []
            
            for _, bakiye_row in bakiye_df.iterrows():
                birlesik_kod = bakiye_row['Birleşik_Kod']
                fatura_miktar = float(bakiye_row['Fatura ve Sevk Edilmemiş Toplam'])
                
                # Eşleşen siparişleri bul
                matching_siparisler = siparis_df[siparis_df['Siparis_Birlesik'] == birlesik_kod]
                matching_grup = siparis_gruplu[siparis_gruplu['Siparis_Birlesik'] == birlesik_kod]
                
                if not matching_grup.empty:
                    toplam_kalan = float(matching_grup.iloc[0]['KALAN_MIKTAR'])
                    
                    # Miktarlar eşleşiyor mu kontrol et
                    if abs(fatura_miktar - toplam_kalan) < 0.001:
                        # Tüm eşleşen satırları ekle
                        for _, siparis_row in matching_siparisler.iterrows():
                            new_row = bakiye_row.to_dict()
                            new_row.update({
                                'SIPARIS_NO': siparis_row['SIPARIS_NO'],
                                'STOK_KODU': siparis_row['STOK_KODU'],
                                'SIPARIS_MIKTARI': float(siparis_row['SIPARIS_MIKTARI']),
                                'KALAN_MIKTAR': float(siparis_row['KALAN_MIKTAR']),
                                'EŞLEŞME_DURUMU': 'TAM_EŞLEŞME'
                            })
                            processed_rows.append(new_row)
                    else:
                        # Kısmi eşleşme
                        new_row = bakiye_row.to_dict()
                        new_row.update({
                            'SIPARIS_NO': '',
                            'STOK_KODU': '',
                            'SIPARIS_MIKTARI': 0,
                            'KALAN_MIKTAR': toplam_kalan,
                            'EŞLEŞME_DURUMU': 'KISMİ_EŞLEŞME'
                        })
                        processed_rows.append(new_row)
                else:
                    # Eşleşme yok
                    new_row = bakiye_row.to_dict()
                    new_row.update({
                        'SIPARIS_NO': '',
                        'STOK_KODU': '',
                        'SIPARIS_MIKTARI': 0,
                        'KALAN_MIKTAR': 0,
                        'EŞLEŞME_DURUMU': 'EŞLEŞME_YOK'
                    })
                    processed_rows.append(new_row)
            
            final_df = pd.DataFrame(processed_rows)

        # 6. ADIM: son.json formatına dönüştür
        with st.spinner("📄 son.json formatı oluşturuluyor..."):
            son_json_data = []
            filtered_count = 0
            
            for _, row in final_df.iterrows():
                depo_kodu = determine_depot_code(row['Sipariş Notu'])
                
                if depo_kodu:
                    son_row = {
                        'Sipariş Notu': str(row['Sipariş Notu']) if pd.notna(row['Sipariş Notu']) else "",
                        'Depo Kodu': depo_kodu,
                        'Ürün Grubu': str(row['Ürün Grubu']) if pd.notna(row['Ürün Grubu']) else "",
                        'Bosch No': str(row['Bosch No']) if pd.notna(row['Bosch No']) else "",
                        'Sütun1': create_sutun1(
                            row['Sipariş Notu'],
                            row['Bosch No'],
                            row.get('SIPARIS_MIKTARI'),
                            row.get('KALAN_MIKTAR')
                        ),
                        'Tahmini Teslim Tarihi': "",
                        'Fatura ve Sevk Edilmemiş Toplam': float(row['Fatura ve Sevk Edilmemiş Toplam']) if pd.notna(row['Fatura ve Sevk Edilmemiş Toplam']) else 0.0,
                        'SIPARIS_MIKTARI': float(row.get('SIPARIS_MIKTARI', 0)),
                        'KALAN_MIKTAR': float(row.get('KALAN_MIKTAR', 0)),
                        'EŞLEŞME_DURUMU': row.get('EŞLEŞME_DURUMU', 'BİLİNMİYOR')
                    }
                    son_json_data.append(son_row)
                else:
                    filtered_count += 1
            
            son_df = pd.DataFrame(son_json_data)
            
            if filtered_count > 0:
                st.warning(f"⚠️ {filtered_count} satır geçersiz depo kodu nedeniyle filtrelendi")
            
            return son_df
            
    except Exception as e:
        st.error(f"❌ İşlem sırasında hata oluştu: {str(e)}")
        return None

def create_excel_file(df):
    """Excel dosyası oluştur"""
    try:
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='BOSCH_Siparisler', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['BOSCH_Siparisler']
            
            # Sütun genişliklerini ayarla
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
            
            # Başlık biçimlendirme
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        output.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"bosch_siparisler_{timestamp}.xlsx"
        
        return output, filename
        
    except Exception as e:
        st.error(f"❌ Excel oluşturma hatası: {str(e)}")
        return None, None

def create_son_json(df):
    """son.json formatında dosya oluştur"""
    try:
        json_data = df.to_dict('records')
        json_str = json.dumps(json_data, indent=2, ensure_ascii=False)
        
        output = io.BytesIO()
        output.write(json_str.encode('utf-8'))
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"son_json_{timestamp}.json"
        
        return output, filename
        
    except Exception as e:
        st.error(f"❌ JSON oluşturma hatası: {str(e)}")
        return None, None

def show_analysis_report(df):
    """Analiz raporunu göster"""
    st.subheader("📊 Analiz Raporu")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.metric("Toplam Kayıt", len(df))
        st.metric("Geçerli Depo Kodlu", len(df[df['Depo Kodu'] != ""]))
        
        if 'EŞLEŞME_DURUMU' in df.columns:
            eslesme_dagilim = df['EŞLEŞME_DURUMU'].value_counts()
            st.write("**Eşleşme Durumu:**")
            st.dataframe(eslesme_dagilim)
    
    with col2:
        st.write("**Depo Dağılımı:**")
        if 'Depo Kodu' in df.columns:
            depo_dagilim = df['Depo Kodu'].value_counts()
            st.dataframe(depo_dagilim)
        
        st.write("**Ürün Grubu Dağılımı:**")
        if 'Ürün Grubu' in df.columns:
            urun_dagilim = df['Ürün Grubu'].value_counts()
            st.dataframe(urun_dagilim)
    
    st.write("**Miktar Analizi:**")
    if 'Fatura ve Sevk Edilmemiş Toplam' in df.columns:
        st.write(f"Toplam Fatura Miktarı: {df['Fatura ve Sevk Edilmemiş Toplam'].sum():,.2f}")
    if 'KALAN_MIKTAR' in df.columns:
        st.write(f"Toplam Kalan Miktar: {df['KALAN_MIKTAR'].sum():,.2f}")

# Sidebar
with st.sidebar:
    st.header("📁 Dosya Yükleme")
    
    bakiye_raporu = st.file_uploader(
        "📊 Bakiye Raporu Excel",
        type=['xlsx', 'xls'],
        key="bakiye"
    )
    
    inbound_excel = st.file_uploader(
        "📦 InBound Excel",
        type=['xlsx', 'xls'],
        key="inbound"
    )
    
    siparis_kalemleri = st.file_uploader(
        "📋 Sipariş Kalemleri Excel",
        type=['xlsx', 'xls'],
        key="siparis"
    )
    
    st.markdown("---")
    
    if st.button("🚀 Verileri İşle", type="primary", use_container_width=True):
        if bakiye_raporu and inbound_excel and siparis_kalemleri:
            with st.spinner("Veriler işleniyor, lütfen bekleyin..."):
                st.session_state.final_df = process_bosch_three_excel(
                    bakiye_raporu, inbound_excel, siparis_kalemleri
                )
        else:
            st.error("Lütfen tüm dosyaları yükleyin!")
    
    st.markdown("---")
    st.header("📋 İşlem Kuralları")
    st.info("""
    1. **Bakiye Raporu:**
       - Bosch No'ya otomatik 3E- eklenir
       - Boşluklar temizlenir
    
    2. **InBound Excel:**
       - BOSCH içeren satırlar filtrelenir
       - DEPO grubuna eklenir
    
    3. **Sipariş Kalemleri:**
       - SIPARIS_NO + STOK_KODU eşleştirilir
       - Miktar kontrolleri yapılır
    
    4. **Çıktı:**
       - Sadece geçerli depo kodları (AAS, DAS, MAS, BAS, EAS)
       - Miktar eşleşmeleri kontrol edilir
    """)

# Ana içerik
if st.session_state.final_df is not None:
    st.success("✅ Veri işleme tamamlandı!")
    
    # Veri önizleme
    st.subheader("📋 İşlenen Veriler")
    st.dataframe(st.session_state.final_df, use_container_width=True)
    
    # Analiz raporu
    show_analysis_report(st.session_state.final_df)
    
    # İndirme butonları
    st.subheader("📥 Çıktı Dosyaları")
    col1, col2 = st.columns(2)
    
    with col1:
        excel_data, excel_name = create_excel_file(st.session_state.final_df)
        if excel_data:
            st.download_button(
                label="📊 Excel Olarak İndir",
                data=excel_data,
                file_name=excel_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    with col2:
        json_data, json_name = create_son_json(st.session_state.final_df)
        if json_data:
            st.download_button(
                label="📄 JSON Olarak İndir",
                data=json_data,
                file_name=json_name,
                mime="application/json",
                use_container_width=True
            )
    
    # Veri inceleme
    st.markdown("---")
    st.subheader("🔍 Veri İnceleme")
    
    if st.checkbox("JSON formatını göster"):
        st.json(st.session_state.final_df.head(5).to_dict('records'))
    
    if st.checkbox("Eşleşme detaylarını göster"):
        st.dataframe(
            st.session_state.final_df[['Sipariş Notu', 'Bosch No', 'Fatura ve Sevk Edilmemiş Toplam', 'KALAN_MIKTAR', 'EŞLEŞME_DURUMU']],
            use_container_width=True
        )

# Sayfa sonu
st.markdown("---")
st.caption("BOSCH Sipariş İşlemleri v3.0 | son.json formatı")
